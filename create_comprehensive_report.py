"""
Create Comprehensive PSX Research Report
Combines all data: stock analysis, news, financial reports into one Excel file
"""
import os
import sys
import json
from pathlib import Path
from datetime import datetime

# Fix encoding
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Try to import pandas and openpyxl
try:
    import pandas as pd
except ImportError:
    os.system("pip install pandas openpyxl --quiet")
    import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Directories
EXPORTS_DIR = Path("exports")
REPORTS_DIR = Path("financial_reports")

def load_stock_analysis():
    """Load the latest stock analysis CSV"""
    csv_files = list(EXPORTS_DIR.glob("stock_analysis_*.csv"))
    if csv_files:
        latest = max(csv_files, key=lambda x: x.stat().st_mtime)
        return pd.read_csv(latest)
    return None

def load_news_data():
    """Load the latest news analysis CSV"""
    csv_files = list(EXPORTS_DIR.glob("news_analysis_*.csv"))
    if csv_files:
        latest = max(csv_files, key=lambda x: x.stat().st_mtime)
        return pd.read_csv(latest)
    return None

def load_financial_reports():
    """Load the financial reports JSON"""
    json_files = list(EXPORTS_DIR.glob("financial_reports_*.json"))
    if json_files:
        latest = max(json_files, key=lambda x: x.stat().st_mtime)
        with open(latest, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def load_announcements():
    """Load announcements from CSV"""
    csv_dirs = list(EXPORTS_DIR.glob("csv_*"))
    if csv_dirs:
        latest_dir = max(csv_dirs, key=lambda x: x.stat().st_mtime)
        ann_file = latest_dir / "announcements.csv"
        if ann_file.exists():
            df = pd.read_csv(ann_file)
            # Filter for top 6 companies
            top6 = ["FFBL", "UBL", "LUCK", "FFC", "MEBL", "PSO"]
            return df[df['symbol'].isin(top6)]
    return None

def create_excel_report():
    """Create comprehensive Excel report"""
    print("=" * 60)
    print("CREATING COMPREHENSIVE PSX RESEARCH REPORT")
    print("=" * 60)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_file = EXPORTS_DIR / f"comprehensive_report_{timestamp}.xlsx"
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(bold=True, size=14)
    buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    hold_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== SHEET 1: EXECUTIVE SUMMARY =====
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    ws1['A1'] = "PSX COMPREHENSIVE RESEARCH REPORT"
    ws1['A1'].font = Font(bold=True, size=18)
    ws1['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M')}"
    ws1['A4'] = "TOP 6 INVESTMENT RECOMMENDATIONS"
    ws1['A4'].font = title_font
    
    # Load stock analysis
    stock_df = load_stock_analysis()
    if stock_df is not None:
        print(f"[OK] Loaded stock analysis: {len(stock_df)} stocks")
        
        # Top 6 summary
        top6_df = stock_df.head(6)
        row = 6
        headers = ["Rank", "Symbol", "Score", "Rating", "Price", "Recommendation"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for i, (_, stock) in enumerate(top6_df.iterrows(), 1):
            row += 1
            ws1.cell(row=row, column=1, value=i)
            ws1.cell(row=row, column=2, value=stock.get('Symbol', ''))
            ws1.cell(row=row, column=3, value=stock.get('Total Score', 0))
            rating = stock.get('Rating', '')
            ws1.cell(row=row, column=4, value=rating)
            ws1.cell(row=row, column=5, value=stock.get('Price', 0))
            
            # Color-coded recommendation
            if 'BUY' in str(rating).upper():
                rec = "STRONG BUY" if stock.get('Total Score', 0) >= 75 else "BUY"
                for col in range(1, 7):
                    ws1.cell(row=row, column=col).fill = buy_fill
            elif 'HOLD' in str(rating).upper():
                rec = "HOLD"
                for col in range(1, 7):
                    ws1.cell(row=row, column=col).fill = hold_fill
            else:
                rec = "AVOID"
                for col in range(1, 7):
                    ws1.cell(row=row, column=col).fill = sell_fill
            
            ws1.cell(row=row, column=6, value=rec)
    
    # ===== SHEET 2: FULL STOCK ANALYSIS =====
    ws2 = wb.create_sheet("Stock Analysis")
    if stock_df is not None:
        for r_idx, row in enumerate(dataframe_to_rows(stock_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
    
    # ===== SHEET 3: FINANCIAL REPORTS =====
    ws3 = wb.create_sheet("Financial Reports")
    ws3['A1'] = "QUARTERLY FINANCIAL REPORTS - TOP 6 COMPANIES"
    ws3['A1'].font = title_font
    
    financial_data = load_financial_reports()
    if financial_data:
        print(f"[OK] Loaded financial reports: {len(financial_data)} companies")
        
        headers = ["Symbol", "Company Name", "Revenue", "Net Profit", "EPS", "Dividend", "Key Figures"]
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for company in financial_data:
            row += 1
            ws3.cell(row=row, column=1, value=company.get('symbol', ''))
            ws3.cell(row=row, column=2, value=company.get('company_name', ''))
            ws3.cell(row=row, column=3, value=company.get('revenue', 'N/A'))
            ws3.cell(row=row, column=4, value=company.get('net_profit', 'N/A'))
            ws3.cell(row=row, column=5, value=company.get('eps', 'N/A'))
            ws3.cell(row=row, column=6, value=company.get('dividend', 'N/A'))
            highlights = "; ".join(company.get('key_highlights', [])[:3])
            ws3.cell(row=row, column=7, value=highlights)
    
    # Add PDF links
    row += 3
    ws3.cell(row=row, column=1, value="PDF REPORT LINKS").font = title_font
    row += 1
    pdf_links = {
        "FFC": "https://dps.psx.com.pk/download/document/264632.pdf",
        "LUCK": "https://dps.psx.com.pk/download/document/264455.pdf",
        "UBL": "https://dps.psx.com.pk/download/document/264343.pdf",
        "MEBL": "https://dps.psx.com.pk/download/document/264198.pdf",
        "PSO": "https://dps.psx.com.pk/download/document/264179.pdf",
        "FFBL": "https://dps.psx.com.pk/download/document/241641.pdf"
    }
    for symbol, link in pdf_links.items():
        row += 1
        ws3.cell(row=row, column=1, value=symbol)
        ws3.cell(row=row, column=2, value=link)
    
    # ===== SHEET 4: NEWS ANALYSIS =====
    ws4 = wb.create_sheet("News Analysis")
    news_df = load_news_data()
    if news_df is not None:
        print(f"[OK] Loaded news: {len(news_df)} articles")
        for r_idx, row in enumerate(dataframe_to_rows(news_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws4.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
    
    # ===== SHEET 5: COMPANY ANNOUNCEMENTS =====
    ws5 = wb.create_sheet("Company Announcements")
    ws5['A1'] = "PSX ANNOUNCEMENTS - TOP 6 COMPANIES"
    ws5['A1'].font = title_font
    
    announcements_df = load_announcements()
    if announcements_df is not None:
        print(f"[OK] Loaded announcements: {len(announcements_df)} items")
        # Keep only relevant columns
        cols = ['symbol', 'announcement_date', 'headline', 'pdf_url', 'announcement_type']
        ann_subset = announcements_df[cols].head(100)
        for r_idx, row in enumerate(dataframe_to_rows(ann_subset, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws5.cell(row=r_idx + 2, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
    
    # ===== SHEET 6: SOURCES & METHODOLOGY =====
    ws6 = wb.create_sheet("Sources & Methodology")
    ws6['A1'] = "DATA SOURCES & METHODOLOGY"
    ws6['A1'].font = title_font
    
    sources = [
        ("Stock Prices", "Pakistan Stock Exchange (dps.psx.com.pk)"),
        ("Company Announcements", "PSX Company Notices"),
        ("Financial Reports", "PSX Quarterly Reports (PDFs)"),
        ("News", "Dawn, Business Recorder, Express Tribune"),
        ("Market Data", "PSX Real-time Data"),
    ]
    
    row = 3
    for source, description in sources:
        ws6.cell(row=row, column=1, value=source).font = Font(bold=True)
        ws6.cell(row=row, column=2, value=description)
        row += 1
    
    row += 2
    ws6.cell(row=row, column=1, value="SCORING METHODOLOGY").font = title_font
    row += 1
    methodology = [
        "Financial Health: 35 points (EPS, Profit Margins, Debt)",
        "Valuation: 25 points (P/E Ratio, Dividend Yield, P/B)",
        "Technical: 20 points (Price Momentum, Volume, RSI)",
        "Sector/Macro: 15 points (Sector Performance, Economic Outlook)",
        "News Sentiment: 5 points (Recent News Analysis)"
    ]
    for item in methodology:
        ws6.cell(row=row, column=1, value=item)
        row += 1
    
    # Adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_file)
    print(f"\n[OK] Report saved: {output_file}")
    
    # Also create a summary CSV
    summary_file = EXPORTS_DIR / f"research_summary_{timestamp}.csv"
    if stock_df is not None:
        top6_summary = stock_df.head(6)[['Symbol', 'Total Score', 'Rating', 'Price', 'EPS', 'P/E Ratio', 'Div Yield']]
        top6_summary.to_csv(summary_file, index=False)
        print(f"[OK] Summary CSV: {summary_file}")
    
    return output_file


if __name__ == "__main__":
    output = create_excel_report()
    print("\n" + "=" * 60)
    print("COMPREHENSIVE REPORT GENERATED SUCCESSFULLY")
    print("=" * 60)
